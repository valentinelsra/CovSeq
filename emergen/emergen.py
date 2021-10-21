#!/usr/local/bin/python3
# -- coding: utf-8 --
# python 3.6

## PRE-REQUIS ##

#pip3 install pandas
#pip3 install openpyxl
#pip3 install requests

## PACKAGES ##

import pandas as pd
import datetime as dt
import openpyxl as opxl
from openpyxl import load_workbook
from openpyxl import Workbook
import urllib
import urllib.request
import requests
import re
import time
import sys

## PARAMETERS ##

#Paths
#path_xlsx='//zisilon03/N02UDC1/NGS COVID/EMERGEN/Copie de YYYY-MM-DD_ACRONYME-PF_semaineXX_emergen_typage_v3.0.6.xlsx'
consensus='En-tête de la séquence dans le fichier fasta'
#Paths à trouver auto avec bash ou python
path_nc= sys.argv[4] + "/LINEAGES/nextclade/nextclade.csv"
path_pg= sys.argv[4] + "/LINEAGES/pangolin/lineage_report.csv"
path_release_nc= sys.argv[4] +"/LINEAGES/nextclade/version.txt"

## SET DATAFRAME W/ DATA TO INSERT INTO XLSX TEMPLATE ##

#Read csv/txt files
nc_df = pd.read_csv(path_nc,error_bad_lines=False, engine="python", sep=";")
pg_df = pd.read_csv(path_pg,error_bad_lines=False, engine="python")
with open(path_release_nc) as f:
    release_nc = f.read()
with open("//zisilon03/N02UDC1/NGS COVID/EMERGEN/version.txt") as f:
    release_xlsx = f.read()

#Filter columns
filtered_nc_df = nc_df.iloc[:,[0,1,3,13,14,15,16,17,18,19,20,21]]
filtered_pg_df = pg_df.iloc[:,[0,1,8,9]]

#Add release column to filtered_nc dataframe
filtered_nc_df.insert(len(filtered_nc_df.columns), 'Version de nextclade', release_nc)

#Change labels of first columns w/ seq consensus name
filtered_nc_df.columns.values[0]=consensus
filtered_pg_df.columns.values[0]=consensus

#Merge dataframes
merged = filtered_nc_df.merge(filtered_pg_df, on=consensus)

#First column -> last column
column_to_move = merged.pop(consensus)
merged.insert(len(merged.columns), consensus, column_to_move)

#Get 'Numéro de prélèvement' from seq consensus names
nums_prelev = []
for seqname in merged[consensus]:
    nums_prelev.append(str(re.findall(r"\d{10}",seqname)[0]))

#Add 'Numéro de prélèvement' column as first
merged.insert(0, 'Numéro de prélèvement', nums_prelev)

#Rename columns w/ template labels
merged.columns.values[len(merged.columns)-4]='Pango lineage'
merged.columns.values[len(merged.columns)-3]='Pangolin version'
merged.columns.values[len(merged.columns)-2]='pangoLEARN version'

#Convert pangoLEARN version column values to date time
#merged.iloc[:,len(merged.columns)-1] = pd.to_datetime(merged.iloc[:,len(merged.columns)-1], format='%Y-%m-%d')

## INSERT DATAFRAME VALUES INTO XLSX TEMPLATE ##

#List of merged dataframe labels
labels = list(merged)

#Load xlsx template and create copy
wb = opxl.load_workbook(sys.argv[1])

#Header row
sheet = wb.worksheets[0]

#Insert merged dataframe values into corresponding columns
for cell in sheet[6]:
    #if cell.value == 'Nomenclature SI-DEP (aide à la saisie du champ JOK3)':
        #print(cell.column)
    if cell.column in range (2,13):
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = 0
    elif cell.value in labels:
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = merged.iloc[i - (cell.row+1)][cell.value]
    elif cell.value == 'Autre information sur le prélèvement (\"CRIBLAGE\" ou \"SEQUENCAGE SYSTEMATIQUE\" par ex)':
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = 'SEQUENCAGE SYSTEMATIQUE'
    elif cell.value == 'Date de résultat du séquençage':
        tmp = str(re.findall(r"\d{6}",sys.argv[3])[0])
        run_date = dt.datetime.strptime(tmp,'%y%m%d').strftime('%Y-%m-%d')
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = run_date
    elif cell.value == 'Technologie de séquençage':
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = 'NGS'
    elif cell.value == 'Longueur de séquençage':
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = 'COMPLET'
    elif cell.value == 'Résultat de la séquence (libre)':
        for i in range(cell.row + 1, len(merged.index) + (cell.row + 1)):
            sheet.cell(row = i, column = cell.column).value = sheet.cell(row = i, column = 44).value
    
#Save modify template as new file
date = str(dt.date.fromtimestamp(time.time()))
num_week = dt.date.fromtimestamp(time.time()).isocalendar()[1]
fname = date + "_CHUBordeaux_semaine" + str(num_week) + "_emergen_typage_" + release_xlsx + '.xlsx'
wb.save(filename = sys.argv[2] + "/" + fname)